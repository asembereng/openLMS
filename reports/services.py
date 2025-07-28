"""
Report generation and export services
"""
from django.utils import timezone
from django.db.models import Sum, Count, Q, Avg
from decimal import Decimal
import json
import io
import csv

from .models import GeneratedReport, ReportExport
from orders.models import Order
from expenses.models import Expense
from customers.models import Customer
from services.models import Service


class ReportGenerationService:
    """Service for generating reports"""
    
    def generate_report(self, template, user, title=None, parameters=None, date_from=None, date_to=None):
        """Generate a report based on template and parameters"""
        parameters = parameters or {}
        
        # Convert date objects to strings for JSON serialization
        serializable_parameters = {}
        for key, value in parameters.items():
            if hasattr(value, 'strftime'):  # Date/datetime object
                serializable_parameters[key] = value.strftime('%Y-%m-%d')
            else:
                serializable_parameters[key] = value
        
        # Create report record
        report = GeneratedReport.objects.create(
            template=template,
            title=title or f"{template.name} - {timezone.now().strftime('%Y-%m-%d %H:%M')}",
            parameters=serializable_parameters,
            date_from=date_from,
            date_to=date_to,
            status='generating',
            generated_by=user,
            expires_at=timezone.now() + timezone.timedelta(hours=24)  # 24-hour retention
        )
        
        try:
            # Generate report data based on template type
            if template.report_type == 'daily_sales':
                data = self._generate_daily_sales_report(date_from, date_to, parameters)
            elif template.report_type == 'monthly_profit':
                data = self._generate_monthly_profit_report(date_from, date_to, parameters)
            elif template.report_type == 'customer_statement':
                data = self._generate_customer_statement(parameters)
            elif template.report_type == 'expense_summary':
                data = self._generate_expense_summary(date_from, date_to, parameters)
            elif template.report_type == 'service_analysis':
                data = self._generate_service_analysis(date_from, date_to, parameters)
            else:
                data = self._generate_custom_report(template, parameters)
            
            # Update report with generated data
            # Normalize any non-JSON-serializable values for JSON serialization
            def _normalize_for_json(obj):
                if isinstance(obj, Decimal):
                    return float(obj)
                if hasattr(obj, 'strftime'):  # datetime/date objects
                    return obj.isoformat()
                if isinstance(obj, dict):
                    return {k: _normalize_for_json(v) for k, v in obj.items()}
                if isinstance(obj, list):
                    return [_normalize_for_json(v) for v in obj]
                return obj
            normalized_data = _normalize_for_json(data)
            report.data = normalized_data
            report.status = 'completed'
            report.generation_time = timezone.now() - report.generated_at
            report.data_size = len(json.dumps(normalized_data).encode('utf-8'))
            report.save()
            
            return report
        
        except Exception as e:
            report.status = 'failed'
            report.error_message = str(e)
            report.save()
            raise
    
    def _generate_daily_sales_report(self, date_from, date_to, parameters):
        """Generate daily sales report"""
        # Handle scope parameter for predefined date ranges
        scope = parameters.get('scope', 'custom')
        status_filter = parameters.get('status_filter', 'completed')
        
        if scope == 'today':
            date_from = date_to = timezone.now().date()
        elif scope == 'week':
            today = timezone.now().date()
            date_from = today - timezone.timedelta(days=today.weekday())
            date_to = date_from + timezone.timedelta(days=6)
        elif scope == 'month':
            today = timezone.now().date()
            date_from = today.replace(day=1)
            next_month = date_from.replace(month=date_from.month + 1) if date_from.month < 12 else date_from.replace(year=date_from.year + 1, month=1)
            date_to = next_month - timezone.timedelta(days=1)
        
        # Build query with status filter
        orders_query = Order.objects.select_related('customer').prefetch_related('lines__service')
        
        if status_filter:
            orders_query = orders_query.filter(status=status_filter)
        else:
            orders_query = orders_query.filter(status='completed')  # Default to completed
        
        if date_from and date_to:
            orders_query = orders_query.filter(created_at__date__range=[date_from, date_to])
        else:
            orders_query = orders_query.filter(created_at__date=timezone.now().date())
        
        orders = orders_query
        
        # Calculate totals
        total_orders = orders.count()
        total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        total_pieces = sum(order.lines.aggregate(total=Sum('pieces'))['total'] or 0 for order in orders)
        
        # Daily breakdown
        daily_data = []
        current_date = date_from or timezone.now().date()
        end_date = date_to or timezone.now().date()
        
        while current_date <= end_date:
            day_orders = orders.filter(created_at__date=current_date)
            day_revenue = day_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
            day_count = day_orders.count()
            
            # Only add dates that have actual orders
            if day_count > 0:
                daily_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'orders': day_count,
                    'revenue': float(day_revenue),
                    'pieces': sum(order.lines.aggregate(total=Sum('pieces'))['total'] or 0 for order in day_orders)
                })
            
            current_date += timezone.timedelta(days=1)
        
        # Top services
        top_services = []
        for order in orders:
            for line in order.lines.all():
                found = False
                for service in top_services:
                    if service['name'] == line.service.name:
                        service['pieces'] += line.pieces
                        service['revenue'] += float(line.line_total)
                        found = True
                        break
                
                if not found:
                    top_services.append({
                        'name': line.service.name,
                        'pieces': line.pieces,
                        'revenue': float(line.line_total)
                    })
        
        top_services.sort(key=lambda x: x['revenue'], reverse=True)
        top_services = top_services[:10]
        
        # Prepare data for table display
        data_for_table = []
        
        # Add daily data to table
        for day in daily_data:
            data_for_table.append({
                'date': day['date'],
                'orders': day['orders'],
                'revenue': day['revenue'],
                'pieces': day['pieces']
            })
        
        # Add top services section
        if top_services:
            data_for_table.append({
                'date': '--- TOP SERVICES ---',
                'orders': '',
                'revenue': '',
                'pieces': ''
            })
            for service in top_services:
                data_for_table.append({
                    'date': service['name'],
                    'orders': '',
                    'revenue': service['revenue'],
                    'pieces': service['pieces']
                })

        return {
            'summary': {
                'Order Count': total_orders,
                'Revenue Amount': float(total_revenue),
                'Pieces Count': total_pieces,
                'Date From': date_from.strftime('%Y-%m-%d') if date_from else None,
                'Date To': date_to.strftime('%Y-%m-%d') if date_to else None
            },
            'daily_data': daily_data,
            'top_services': top_services,
            'data': data_for_table,  # For template table display
            'generated_at': timezone.now().isoformat()
        }
    
    def _generate_monthly_profit_report(self, date_from, date_to, parameters):
        """Generate monthly profit report"""
        # Revenue from completed orders
        orders = Order.objects.filter(status='completed')
        if date_from and date_to:
            orders = orders.filter(created_at__date__range=[date_from, date_to])
        
        total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        
        # Expenses
        expenses = Expense.objects.filter(is_approved=True)
        if date_from and date_to:
            expenses = expenses.filter(expense_date__range=[date_from, date_to])
        
        total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        # Calculate profit
        gross_profit = total_revenue - total_expenses
        profit_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
        
        # Monthly breakdown
        monthly_data = []
        if date_from and date_to:
            current_date = date_from.replace(day=1)
            end_date = date_to
            
            while current_date <= end_date:
                next_month = current_date.replace(month=current_date.month + 1) if current_date.month < 12 else current_date.replace(year=current_date.year + 1, month=1)
                
                month_orders = orders.filter(
                    created_at__date__gte=current_date,
                    created_at__date__lt=next_month
                )
                month_expenses = expenses.filter(
                    expense_date__gte=current_date,
                    expense_date__lt=next_month
                )
                
                month_revenue = month_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
                month_expense = month_expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                month_profit = month_revenue - month_expense
                
                # Only add months that have actual data
                if month_revenue > 0 or month_expense > 0:
                    monthly_data.append({
                        'month': current_date.strftime('%Y-%m'),
                        'month_name': current_date.strftime('%B %Y'),
                        'revenue': float(month_revenue),
                        'expenses': float(month_expense),
                        'profit': float(month_profit)
                    })
                
                current_date = next_month
        
        # Prepare data for table display
        data_for_table = []
        
        # Add monthly data to table
        for month in monthly_data:
            data_for_table.append({
                'month': month['month_name'],
                'revenue': month['revenue'],
                'expenses': month['expenses'],
                'profit': month['profit']
            })

        return {
            'summary': {
                'Revenue Amount': float(total_revenue),
                'Expense Amount': float(total_expenses),
                'Profit Amount': float(gross_profit),
                'Profit Margin %': float(profit_margin),
                'Date From': date_from.strftime('%Y-%m-%d') if date_from else None,
                'Date To': date_to.strftime('%Y-%m-%d') if date_to else None
            },
            'monthly_data': monthly_data,
            'data': data_for_table,  # For template table display
            'generated_at': timezone.now().isoformat()
        }
    
    def _generate_customer_statement(self, parameters):
        """Generate customer statement"""
        customer_id = parameters.get('customer_id')
        if not customer_id:
            raise ValueError("Customer ID is required for customer statement")
        
        customer = Customer.objects.get(id=customer_id)
        orders = Order.objects.filter(customer=customer).select_related('customer').prefetch_related('lines__service')
        
        # Calculate totals
        total_orders = orders.count()
        total_amount = orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        completed_orders = orders.filter(status='completed').count()
        pending_orders = orders.filter(status__in=['pending', 'processing']).count()
        
        # Order history
        order_history = []
        for order in orders.order_by('-created_at')[:20]:  # Last 20 orders
            order_history.append({
                'order_number': order.order_number,
                'date': order.created_at.strftime('%Y-%m-%d'),
                'status': order.status,
                'total_amount': float(order.total_amount),
                'pieces': order.lines.aggregate(total=Sum('pieces'))['total'] or 0
            })
        
        # Prepare data for table display
        data_for_table = []
        
        # Add order history to table
        for order in order_history:
            data_for_table.append({
                'order_number': order['order_number'],
                'date': order['date'],
                'status': order['status'],
                'total_amount': order['total_amount'],
                'pieces': order['pieces']
            })

        return {
            'customer': {
                'name': customer.name,
                'phone': customer.phone,
                'email': customer.email,
                'address': customer.address
            },
            'summary': {
                'Order Count': total_orders,
                'Amount Total': float(total_amount),
                'Orders Completed': completed_orders,
                'Orders Pending': pending_orders
            },
            'order_history': order_history,
            'data': data_for_table,  # For template table display
            'generated_at': timezone.now().isoformat()
        }
    
    def _generate_expense_summary(self, date_from, date_to, parameters):
        """Generate expense summary report"""
        expenses = Expense.objects.select_related('category', 'created_by')
        
        if date_from and date_to:
            expenses = expenses.filter(expense_date__range=[date_from, date_to])
        
        # Calculate totals
        total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        approved_expenses = expenses.filter(is_approved=True).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        pending_expenses = total_expenses - approved_expenses
        
        # Category breakdown
        category_breakdown = list(
            expenses.values('category__name', 'category__color')
            .annotate(total=Sum('amount'), count=Count('id'))
            .order_by('-total')
        )
        
        # Prepare data for table display
        data_for_table = []
        
        # Add category breakdown to table
        for category in category_breakdown:
            data_for_table.append({
                'category': category['category__name'] or 'Uncategorized',
                'total_amount': category['total'],
                'expense_count': category['count']
            })

        return {
            'summary': {
                'Expense Amount': float(total_expenses),
                'Approved Amount': float(approved_expenses),
                'Pending Amount': float(pending_expenses),
                'Expense Count': expenses.count(),
                'Date From': date_from.strftime('%Y-%m-%d') if date_from else None,
                'Date To': date_to.strftime('%Y-%m-%d') if date_to else None
            },
            'category_breakdown': category_breakdown,
            'data': data_for_table,  # For template table display
            'generated_at': timezone.now().isoformat()
        }
    
    def _generate_service_analysis(self, date_from, date_to, parameters):
        """Generate service analysis report"""
        from orders.models import OrderLine
        
        order_lines = OrderLine.objects.select_related('service', 'order')
        
        if date_from and date_to:
            order_lines = order_lines.filter(order__created_at__date__range=[date_from, date_to])
        
        # Service performance
        service_performance = list(
            order_lines.values('service__name', 'service__category__name')
            .annotate(
                total_pieces=Sum('pieces'),
                total_revenue=Sum('line_total'),
                order_count=Count('order', distinct=True)
            )
            .order_by('-total_revenue')
        )
        
        # Prepare data for table display
        data_for_table = []
        
        # Add service performance to table
        for service in service_performance:
            data_for_table.append({
                'service_name': service['service__name'],
                'category': service['service__category__name'] or 'Uncategorized',
                'total_pieces': service['total_pieces'],
                'total_revenue': service['total_revenue'],
                'order_count': service['order_count']
            })

        return {
            'summary': {
                'Service Count': len(service_performance),
                'Date From': date_from.strftime('%Y-%m-%d') if date_from else None,
                'Date To': date_to.strftime('%Y-%m-%d') if date_to else None
            },
            'service_performance': service_performance,
            'data': data_for_table,  # For template table display
            'generated_at': timezone.now().isoformat()
        }
    
    def _generate_custom_report(self, template, parameters):
        """Generate custom report based on template configuration"""
        try:
            config = template.config or {}
            report_data = []
            summary_data = {}
            
            # Get the data source from config
            data_source = config.get('data_source', 'orders')
            date_field = config.get('date_field', 'created_at')
            filters = config.get('filters', {})
            columns = config.get('columns', [])
            aggregations = config.get('aggregations', {})
            
            # Apply date filtering from parameters
            date_from = parameters.get('date_from')
            date_to = parameters.get('date_to')
            
            # Convert string dates back to date objects if needed
            if isinstance(date_from, str):
                try:
                    from datetime import datetime
                    date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                except:
                    date_from = None
            
            if isinstance(date_to, str):
                try:
                    from datetime import datetime
                    date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                except:
                    date_to = None
            
            # Build the base queryset based on data source
            if data_source == 'orders':
                queryset = Order.objects.select_related('customer', 'payment_method')
                if date_from and date_to:
                    queryset = queryset.filter(**{f"{date_field}__date__range": [date_from, date_to]})
                
                # Apply additional filters from config
                for filter_field, filter_value in filters.items():
                    if filter_value:
                        queryset = queryset.filter(**{filter_field: filter_value})
                
                # Get the data
                for order in queryset.order_by('-created_at')[:1000]:  # Limit to 1000 records
                    row = {}
                    for column in columns:
                        field_name = column.get('field')
                        display_name = column.get('name', field_name)
                        
                        if field_name == 'order_number':
                            row[display_name] = order.order_number
                        elif field_name == 'customer_name':
                            row[display_name] = order.customer.name
                        elif field_name == 'customer_phone':
                            row[display_name] = order.customer.phone
                        elif field_name == 'status':
                            row[display_name] = order.get_status_display()
                        elif field_name == 'total_amount':
                            row[display_name] = float(order.total_amount)
                        elif field_name == 'payment_method':
                            row[display_name] = order.payment_method.name
                        elif field_name == 'created_at':
                            row[display_name] = order.created_at.strftime('%Y-%m-%d %H:%M')
                        elif field_name == 'total_pieces':
                            row[display_name] = order.total_pieces
                        else:
                            # Try to get the field value dynamically
                            try:
                                value = getattr(order, field_name, 'N/A')
                                row[display_name] = value
                            except:
                                row[display_name] = 'N/A'
                    
                    report_data.append(row)
                
                # Calculate aggregations
                if aggregations:
                    if 'total_revenue' in aggregations:
                        total_revenue = queryset.aggregate(total=Sum('total_amount'))['total'] or 0
                        summary_data['Total Revenue'] = float(total_revenue)
                    
                    if 'order_count' in aggregations:
                        summary_data['Order Count'] = queryset.count()
                    
                    if 'avg_order_value' in aggregations:
                        avg_value = queryset.aggregate(avg=Avg('total_amount'))['avg'] or 0
                        summary_data['Average Order Value'] = float(avg_value)
            
            elif data_source == 'customers':
                from customers.models import Customer
                queryset = Customer.objects.all()
                
                # Apply filters
                for filter_field, filter_value in filters.items():
                    if filter_value:
                        queryset = queryset.filter(**{filter_field: filter_value})
                
                for customer in queryset.order_by('name')[:1000]:
                    row = {}
                    for column in columns:
                        field_name = column.get('field')
                        display_name = column.get('name', field_name)
                        
                        if field_name == 'name':
                            row[display_name] = customer.name
                        elif field_name == 'phone':
                            row[display_name] = customer.phone
                        elif field_name == 'email':
                            row[display_name] = customer.email or 'N/A'
                        elif field_name == 'total_orders':
                            row[display_name] = customer.orders.count()
                        elif field_name == 'total_spent':
                            total_spent = customer.orders.aggregate(total=Sum('total_amount'))['total'] or 0
                            row[display_name] = float(total_spent)
                        elif field_name == 'last_order_date':
                            last_order = customer.orders.order_by('-created_at').first()
                            row[display_name] = last_order.created_at.strftime('%Y-%m-%d') if last_order else 'Never'
                        else:
                            try:
                                value = getattr(customer, field_name, 'N/A')
                                row[display_name] = value
                            except:
                                row[display_name] = 'N/A'
                    
                    report_data.append(row)
                
                # Customer aggregations
                if aggregations:
                    if 'customer_count' in aggregations:
                        summary_data['Customer Count'] = queryset.count()
                    
                    if 'total_revenue' in aggregations:
                        total_revenue = Order.objects.filter(customer__in=queryset).aggregate(total=Sum('total_amount'))['total'] or 0
                        summary_data['Total Revenue'] = float(total_revenue)
            
            elif data_source == 'expenses':
                queryset = Expense.objects.select_related('category', 'created_by')
                if date_from and date_to:
                    queryset = queryset.filter(expense_date__range=[date_from, date_to])
                
                # Apply filters
                for filter_field, filter_value in filters.items():
                    if filter_value:
                        queryset = queryset.filter(**{filter_field: filter_value})
                
                for expense in queryset.order_by('-expense_date')[:1000]:
                    row = {}
                    for column in columns:
                        field_name = column.get('field')
                        display_name = column.get('name', field_name)
                        
                        if field_name == 'description':
                            row[display_name] = expense.description
                        elif field_name == 'amount':
                            row[display_name] = float(expense.amount)
                        elif field_name == 'category':
                            row[display_name] = expense.category.name if expense.category else 'Uncategorized'
                        elif field_name == 'is_approved':
                            row[display_name] = 'Approved' if expense.is_approved else 'Pending'
                        elif field_name == 'expense_date':
                            row[display_name] = expense.expense_date.strftime('%Y-%m-%d')
                        elif field_name == 'created_by':
                            row[display_name] = expense.created_by.get_full_name() or expense.created_by.username
                        else:
                            try:
                                value = getattr(expense, field_name, 'N/A')
                                row[display_name] = value
                            except:
                                row[display_name] = 'N/A'
                    
                    report_data.append(row)
                
                # Expense aggregations
                if aggregations:
                    if 'total_expenses' in aggregations:
                        total_expenses = queryset.aggregate(total=Sum('amount'))['total'] or 0
                        summary_data['Total Expenses'] = float(total_expenses)
                    
                    if 'expense_count' in aggregations:
                        summary_data['Expense Count'] = queryset.count()
                    
                    if 'approved_expenses' in aggregations:
                        approved_total = queryset.filter(is_approved=True).aggregate(total=Sum('amount'))['total'] or 0
                        summary_data['Approved Expenses'] = float(approved_total)
            
            else:
                # Fallback for unknown data sources
                report_data = [{'info': f'Unknown data source: {data_source}'}]
                summary_data = {'Error': 'Invalid data source configuration'}
            
            # Add metadata to summary
            summary_data['Records Returned'] = len(report_data)
            summary_data['Data Source'] = data_source.title()
            if date_from:
                summary_data['Date From'] = date_from.strftime('%Y-%m-%d') if hasattr(date_from, 'strftime') else str(date_from)
            if date_to:
                summary_data['Date To'] = date_to.strftime('%Y-%m-%d') if hasattr(date_to, 'strftime') else str(date_to)
            
            return {
                'summary': summary_data,
                'report_data': report_data,
                'data': report_data,  # For template table display
                'config': config,
                'generated_at': timezone.now().isoformat()
            }
            
        except Exception as e:
            return {
                'summary': {'Error': str(e)},
                'report_data': [],
                'data': [{'error': f'Failed to generate custom report: {str(e)}'}],
                'generated_at': timezone.now().isoformat()
            }


class ReportExportService:
    """Service for exporting reports to various formats"""
    
    def export_report(self, report, format_type, user):
        """Export a report to the specified format and return HttpResponse for download"""
        if format_type == 'csv':
            return self._export_to_csv(report, user)
        elif format_type == 'json':
            return self._export_to_json(report, user)
        elif format_type == 'pdf':
            return self._export_to_pdf(report, user)
        elif format_type == 'excel':
            return self._export_to_excel(report, user)
        else:
            raise ValueError(f"Unsupported export format: {format_type}")
    
    def create_export_record(self, report, format_type, user):
        """Create an export record for API responses (without file content)"""
        file_name = f"{report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{format_type}"
        
        export = ReportExport.objects.create(
            report=report,
            format=format_type,
            file_name=file_name,
            file_size=0,  # Placeholder - file is generated on demand
            exported_by=user
        )
        
        return export
    
    def _export_to_csv(self, report, user):
        """Export report to CSV format"""
        from django.http import HttpResponse
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Report:', report.title])
        writer.writerow(['Generated:', report.generated_at.strftime('%Y-%m-%d %H:%M')])
        writer.writerow(['Generated By:', user.get_full_name() or user.username])
        writer.writerow([])  # Empty row
        
        # Write data based on report structure
        data = report.data
        if isinstance(data, str):
            data = json.loads(data)
        
        # Write summary section
        if 'summary' in data and data['summary']:
            writer.writerow(['=== SUMMARY ==='])
            for key, value in data['summary'].items():
                writer.writerow([key.replace('_', ' ').title(), str(value)])
            writer.writerow([])
        
        # Write main data section
        if 'data' in data and data['data']:
            # Get the headers from the first row
            if len(data['data']) > 0:
                headers = list(data['data'][0].keys())
                writer.writerow(['=== DATA ==='])
                writer.writerow(headers)
                
                for row in data['data']:
                    writer.writerow([str(row.get(header, '')) for header in headers])
        
        # If no main data, try other data structures
        elif 'daily_data' in data and data['daily_data']:
            writer.writerow(['=== DAILY DATA ==='])
            writer.writerow(['Date', 'Orders', 'Revenue', 'Pieces'])
            for item in data['daily_data']:
                writer.writerow([
                    item.get('date', ''), 
                    item.get('orders', 0), 
                    item.get('revenue', 0), 
                    item.get('pieces', 0)
                ])
        
        elif 'service_performance' in data and data['service_performance']:
            writer.writerow(['=== SERVICE PERFORMANCE ==='])
            writer.writerow(['Service Name', 'Category', 'Total Pieces', 'Total Revenue', 'Order Count'])
            for item in data['service_performance']:
                writer.writerow([
                    item.get('service__name', ''),
                    item.get('service__category__name', 'Uncategorized'),
                    item.get('total_pieces', 0),
                    item.get('total_revenue', 0),
                    item.get('order_count', 0)
                ])
        
        # Create the HTTP response
        csv_content = output.getvalue()
        file_name = f"{report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        # Create export record
        export = ReportExport.objects.create(
            report=report,
            format='csv',
            file_name=file_name,
            file_size=len(csv_content.encode('utf-8')),
            exported_by=user
        )
        
        # Return the file as HTTP response
        response = HttpResponse(csv_content, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response
    
    def _export_to_json(self, report, user):
        """Export report to JSON format"""
        from django.http import HttpResponse
        
        # Create formatted JSON content
        export_data = {
            'report_info': {
                'title': report.title,
                'generated_at': report.generated_at.isoformat(),
                'generated_by': user.get_full_name() or user.username,
                'template': report.template.name,
                'report_type': report.template.get_report_type_display()
            },
            'report_data': report.data
        }
        
        json_content = json.dumps(export_data, indent=2, default=str)
        file_name = f"{report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        # Create export record
        export = ReportExport.objects.create(
            report=report,
            format='json',
            file_name=file_name,
            file_size=len(json_content.encode('utf-8')),
            exported_by=user
        )
        
        # Return the file as HTTP response
        response = HttpResponse(json_content, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response
    
    def _export_to_pdf(self, report, user):
        """Export report to PDF format"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from django.http import HttpResponse
            
            # Create PDF content
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center
            )
            story.append(Paragraph(report.title, title_style))
            story.append(Spacer(1, 12))
            
            # Report info
            info_style = styles['Normal']
            story.append(Paragraph(f"<b>Generated:</b> {report.generated_at.strftime('%Y-%m-%d %H:%M')}", info_style))
            story.append(Paragraph(f"<b>Generated By:</b> {user.get_full_name() or user.username}", info_style))
            story.append(Paragraph(f"<b>Report Type:</b> {report.template.get_report_type_display()}", info_style))
            story.append(Spacer(1, 20))
            
            # Process report data
            data = report.data
            if isinstance(data, str):
                data = json.loads(data)
            
            # Summary section
            if 'summary' in data and data['summary']:
                story.append(Paragraph("Summary", styles['Heading2']))
                summary_data = [['Metric', 'Value']]
                for key, value in data['summary'].items():
                    summary_data.append([key.replace('_', ' ').title(), str(value)])
                
                summary_table = Table(summary_data)
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(summary_table)
                story.append(Spacer(1, 20))
            
            # Data table
            if 'data' in data and data['data'] and len(data['data']) > 0:
                story.append(Paragraph("Detailed Data", styles['Heading2']))
                
                # Get headers and limit data for PDF
                headers = list(data['data'][0].keys())
                table_data = [headers]
                
                # Limit to first 50 rows for PDF
                for row in data['data'][:50]:
                    table_data.append([str(row.get(header, '')) for header in headers])
                
                if len(data['data']) > 50:
                    story.append(Paragraph(f"<i>Showing first 50 of {len(data['data'])} records</i>", styles['Italic']))
                
                data_table = Table(table_data)
                data_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                story.append(data_table)
            
            # Build PDF
            doc.build(story)
            pdf_content = buffer.getvalue()
            buffer.close()
            
            file_name = f"{report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Create export record
            export = ReportExport.objects.create(
                report=report,
                format='pdf',
                file_name=file_name,
                file_size=len(pdf_content),
                exported_by=user
            )
            
            # Return the file as HTTP response
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
            
        except ImportError:
            # Fallback if reportlab is not installed
            from django.http import HttpResponse
            
            file_name = f"{report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.txt"
            
            # Create a text file instead
            content = f"Report: {report.title}\n"
            content += f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M')}\n"
            content += f"Generated By: {user.get_full_name() or user.username}\n\n"
            content += "PDF export requires reportlab library.\n"
            content += "Install with: pip install reportlab\n\n"
            content += f"Report Data:\n{json.dumps(report.data, indent=2, default=str)}"
            
            # Create export record
            export = ReportExport.objects.create(
                report=report,
                format='pdf',
                file_name=file_name,
                file_size=len(content.encode('utf-8')),
                exported_by=user
            )
            
            response = HttpResponse(content, content_type='text/plain')
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
    
    def _export_to_excel(self, report, user):
        """Export report to Excel format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Report Data"
                
                # Add report info
                ws['A1'] = "Report:"
                ws['B1'] = report.title
                ws['A2'] = "Generated:"
                ws['B2'] = report.generated_at.strftime('%Y-%m-%d %H:%M')
                ws['A3'] = "Generated By:"
                ws['B3'] = user.get_full_name() or user.username
                
                # Style the header
                header_font = Font(bold=True)
                for cell_ref in ['A1', 'A2', 'A3']:
                    cell = ws[cell_ref]
                    if cell is not None:
                        cell.font = header_font
                
                # Process report data
                data = report.data
                if isinstance(data, str):
                    data = json.loads(data)
                
                current_row = 5
                
                # Summary section
                if 'summary' in data and data['summary']:
                    summary_cell = ws[f'A{current_row}']
                    if summary_cell is not None:
                        summary_cell.value = "SUMMARY"
                        summary_cell.font = Font(bold=True, size=14)
                    current_row += 1
                    
                    for key, value in data['summary'].items():
                        key_cell = ws[f'A{current_row}']
                        value_cell = ws[f'B{current_row}']
                        if key_cell is not None:
                            key_cell.value = key.replace('_', ' ').title()
                        if value_cell is not None:
                            value_cell.value = str(value)
                        current_row += 1
                    
                    current_row += 2
                
                # Data section
                if 'data' in data and data['data'] and len(data['data']) > 0:
                    data_header_cell = ws[f'A{current_row}']
                    if data_header_cell is not None:
                        data_header_cell.value = "DETAILED DATA"
                        data_header_cell.font = Font(bold=True, size=14)
                    current_row += 1
                    
                    # Headers
                    headers = list(data['data'][0].keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col)
                        if cell is not None:
                            cell.value = header.replace('_', ' ').title()
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    
                    current_row += 1
                    
                    # Data rows
                    for row_data in data['data']:
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=current_row, column=col)
                            if cell is not None:
                                value = row_data.get(header, '')
                                cell.value = str(value)
                        current_row += 1
                
                # Auto-adjust column widths
                if hasattr(ws, 'columns'):
                    for column in ws.columns:
                        if column and len(column) > 0:
                            max_length = 0
                            column_letter = column[0].column_letter if hasattr(column[0], 'column_letter') else 'A'
                            for cell in column:
                                try:
                                    if cell and cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            if hasattr(ws, 'column_dimensions'):
                                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()
            
            file_name = f"{report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Create export record
            export = ReportExport.objects.create(
                report=report,
                format='excel',
                file_name=file_name,
                file_size=len(excel_content),
                exported_by=user
            )
            
            # Return the file as HTTP response
            response = HttpResponse(
                excel_content, 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
            
        except ImportError:
            # Fallback if openpyxl is not installed
            from django.http import HttpResponse
            
            # Create CSV instead
            return self._export_to_csv(report, user)
